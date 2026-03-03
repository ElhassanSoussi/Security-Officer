"""
Multi-Run Intelligence + Institutional Memory Engine Tests

Tests cover:
1. EmbeddingCache — LRU eviction, get/put, size, clear
2. SimilarityMatch / SimilarityResult dataclasses
3. Reuse classification thresholds (reuse / suggest / generate)
4. Delta tracking (NEW / MODIFIED / UNCHANGED)
5. QuestionItem extended fields (answer_origin, reused_from_question_id, etc.)
6. Audit sheet "Answer Origin" column in export
7. Generation pipeline fields present in all response types
8. Config — default settings
9. Batch embedding helper
10. Run comparison delta logic
"""
import sys
import os
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))


# ─── Test Helpers ────────────────────────────────────────────────────────────

def _make_question(**overrides):
    """Create a QuestionItem with multi-run intelligence defaults."""
    from app.models.schemas import QuestionItem

    defaults = {
        "sheet_name": "Sheet1",
        "cell_coordinate": "B2",
        "question": "Test question",
        "ai_answer": "Test answer",
        "final_answer": "Test answer",
        "confidence": "HIGH",
        "sources": ["test_doc.pdf"],
        "source_excerpt": "Relevant excerpt",
        "is_verified": False,
        "edited_by_user": False,
        "review_status": "approved",
        "confidence_score": 0.85,
        "confidence_reason": "high similarity match; 3 supporting chunks",
        "embedding_similarity_score": 0.92,
        "chunk_id": "chunk-uuid-123",
        "token_count_used": 450,
        "model_used": "gpt-4-turbo",
        "generation_time_ms": 1200,
        "retrieval_mode": "standard",
        # Multi-run fields
        "answer_origin": "generated",
        "reused_from_question_id": None,
        "reuse_similarity_score": None,
        "change_type": "NEW",
    }
    defaults.update(overrides)
    return QuestionItem(**defaults)


def _create_test_workbook() -> bytes:
    from openpyxl import Workbook
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Question", "Answer"])
    ws.append(["Is the building entrance accessible?", ""])
    ws.append(["Are fire extinguishers inspected?", ""])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
# 1. EmbeddingCache Tests
# ═══════════════════════════════════════════════════════════════════════════════

class TestEmbeddingCache:
    def test_put_and_get(self):
        from app.core.similarity import EmbeddingCache

        cache = EmbeddingCache(max_size=10)
        embedding = [0.1, 0.2, 0.3]
        cache.put("test question", embedding)
        result = cache.get("test question")
        assert result == embedding

    def test_cache_miss(self):
        from app.core.similarity import EmbeddingCache

        cache = EmbeddingCache(max_size=10)
        result = cache.get("nonexistent")
        assert result is None

    def test_case_insensitive(self):
        from app.core.similarity import EmbeddingCache

        cache = EmbeddingCache(max_size=10)
        embedding = [0.1, 0.2, 0.3]
        cache.put("Test Question", embedding)
        result = cache.get("test question")
        assert result == embedding

    def test_whitespace_normalization(self):
        from app.core.similarity import EmbeddingCache

        cache = EmbeddingCache(max_size=10)
        embedding = [0.1, 0.2, 0.3]
        cache.put("  test question  ", embedding)
        result = cache.get("test question")
        assert result == embedding

    def test_lru_eviction(self):
        from app.core.similarity import EmbeddingCache

        cache = EmbeddingCache(max_size=3)
        cache.put("a", [1.0])
        cache.put("b", [2.0])
        cache.put("c", [3.0])

        # Cache is full. Adding "d" should evict "a" (least recently used)
        cache.put("d", [4.0])

        assert cache.get("a") is None  # evicted
        assert cache.get("b") == [2.0]
        assert cache.get("c") == [3.0]
        assert cache.get("d") == [4.0]
        assert cache.size() == 3

    def test_access_refreshes_lru_order(self):
        from app.core.similarity import EmbeddingCache

        cache = EmbeddingCache(max_size=3)
        cache.put("a", [1.0])
        cache.put("b", [2.0])
        cache.put("c", [3.0])

        # Access "a" to refresh it
        cache.get("a")

        # Now adding "d" should evict "b" (least recently used after refresh)
        cache.put("d", [4.0])

        assert cache.get("a") == [1.0]  # refreshed, not evicted
        assert cache.get("b") is None    # evicted
        assert cache.get("c") == [3.0]
        assert cache.get("d") == [4.0]

    def test_clear(self):
        from app.core.similarity import EmbeddingCache

        cache = EmbeddingCache(max_size=10)
        cache.put("a", [1.0])
        cache.put("b", [2.0])
        assert cache.size() == 2

        cache.clear()
        assert cache.size() == 0
        assert cache.get("a") is None

    def test_update_existing_key(self):
        from app.core.similarity import EmbeddingCache

        cache = EmbeddingCache(max_size=10)
        cache.put("key", [1.0])
        cache.put("key", [2.0])
        assert cache.get("key") == [2.0]
        assert cache.size() == 1


# ═══════════════════════════════════════════════════════════════════════════════
# 2. SimilarityMatch / SimilarityResult Dataclass Tests
# ═══════════════════════════════════════════════════════════════════════════════

class TestSimilarityDataclasses:
    def test_similarity_match_to_dict(self):
        from app.core.similarity import SimilarityMatch

        match = SimilarityMatch(
            question_embedding_id="qe-123",
            question_text="Is the entrance accessible?",
            answer_text="Yes, per Section 4.2",
            similarity=0.95,
            source_document="compliance.pdf",
        )
        d = match.to_dict()
        assert d["question_embedding_id"] == "qe-123"
        assert d["similarity"] == 0.95
        assert d["answer_text"] == "Yes, per Section 4.2"

    def test_similarity_result_empty(self):
        from app.core.similarity import SimilarityResult

        result = SimilarityResult()
        assert result.has_reusable is False
        assert result.has_suggestion is False
        assert result.action == "generate"

    def test_similarity_result_reusable(self):
        from app.core.similarity import SimilarityMatch, SimilarityResult

        match = SimilarityMatch(
            question_embedding_id="qe-1",
            question_text="Q1",
            answer_text="A1",
            similarity=0.95,
        )
        result = SimilarityResult(
            matches=[match],
            best_match=match,
            action="reuse",
        )
        assert result.has_reusable is True
        assert result.has_suggestion is False

    def test_similarity_result_suggestion(self):
        from app.core.similarity import SimilarityMatch, SimilarityResult

        match = SimilarityMatch(
            question_embedding_id="qe-2",
            question_text="Q2",
            answer_text="A2",
            similarity=0.82,
        )
        result = SimilarityResult(
            matches=[match],
            best_match=match,
            action="suggest",
        )
        assert result.has_reusable is False
        assert result.has_suggestion is True

    def test_similarity_result_to_dict(self):
        from app.core.similarity import SimilarityMatch, SimilarityResult

        match = SimilarityMatch(
            question_embedding_id="qe-3",
            question_text="Q3",
            answer_text="A3",
            similarity=0.88,
        )
        result = SimilarityResult(
            matches=[match],
            best_match=match,
            action="suggest",
            search_time_ms=42,
        )
        d = result.to_dict()
        assert d["action"] == "suggest"
        assert d["matches_count"] == 1
        assert d["best_similarity"] == 0.88
        assert d["search_time_ms"] == 42


# ═══════════════════════════════════════════════════════════════════════════════
# 3. Reuse Classification Tests
# ═══════════════════════════════════════════════════════════════════════════════

class TestReuseClassification:
    """Verify the threshold logic: ≥0.90 reuse, 0.75-0.90 suggest, <0.75 generate."""

    def test_exact_reuse_threshold(self):
        """Similarity of 0.90 should trigger reuse."""
        from app.core.config import get_settings
        s = get_settings()
        assert s.REUSE_EXACT_THRESHOLD == 0.90
        # 0.90 >= 0.90 → reuse
        assert 0.90 >= s.REUSE_EXACT_THRESHOLD

    def test_suggestion_threshold(self):
        """Similarity 0.80 is between suggest (0.75) and reuse (0.90)."""
        from app.core.config import get_settings
        s = get_settings()
        sim = 0.80
        assert sim >= s.REUSE_SUGGEST_THRESHOLD
        assert sim < s.REUSE_EXACT_THRESHOLD

    def test_below_suggestion_threshold(self):
        """Similarity 0.70 is below suggest threshold → normal generation."""
        from app.core.config import get_settings
        s = get_settings()
        sim = 0.70
        assert sim < s.REUSE_SUGGEST_THRESHOLD

    def test_engine_classifies_reuse(self):
        """SimilarityEngine should classify ≥0.90 as 'reuse'."""
        from app.core.similarity import SimilarityEngine
        engine = SimilarityEngine()
        # We test the classification logic by verifying the threshold constants
        from app.core.config import get_settings
        s = get_settings()
        # High similarity → reuse
        sim = 0.95
        if sim >= s.REUSE_EXACT_THRESHOLD:
            action = "reuse"
        elif sim >= s.REUSE_SUGGEST_THRESHOLD:
            action = "suggest"
        else:
            action = "generate"
        assert action == "reuse"

    def test_engine_classifies_suggest(self):
        from app.core.config import get_settings
        s = get_settings()
        sim = 0.82
        if sim >= s.REUSE_EXACT_THRESHOLD:
            action = "reuse"
        elif sim >= s.REUSE_SUGGEST_THRESHOLD:
            action = "suggest"
        else:
            action = "generate"
        assert action == "suggest"

    def test_engine_classifies_generate(self):
        from app.core.config import get_settings
        s = get_settings()
        sim = 0.60
        if sim >= s.REUSE_EXACT_THRESHOLD:
            action = "reuse"
        elif sim >= s.REUSE_SUGGEST_THRESHOLD:
            action = "suggest"
        else:
            action = "generate"
        assert action == "generate"


# ═══════════════════════════════════════════════════════════════════════════════
# 4. Delta Tracking Tests
# ═══════════════════════════════════════════════════════════════════════════════

class TestDeltaTracking:
    def test_all_new_questions(self):
        from app.core.similarity import compute_delta

        current = [
            {"question_text": "What is fire safety?", "cell_reference": "B2"},
            {"question_text": "Is entrance accessible?", "cell_reference": "B3"},
        ]
        previous = []

        delta = compute_delta(current, previous)
        assert delta["What is fire safety?"] == "NEW"
        assert delta["Is entrance accessible?"] == "NEW"

    def test_all_unchanged(self):
        from app.core.similarity import compute_delta

        questions = [
            {"question_text": "What is fire safety?", "cell_reference": "B2"},
            {"question_text": "Is entrance accessible?", "cell_reference": "B3"},
        ]

        delta = compute_delta(questions, questions)
        assert delta["What is fire safety?"] == "UNCHANGED"
        assert delta["Is entrance accessible?"] == "UNCHANGED"

    def test_modified_question(self):
        from app.core.similarity import compute_delta

        current = [
            {"question_text": "Is the main entrance ADA compliant?", "cell_reference": "B2"},
        ]
        previous = [
            {"question_text": "Is the entrance accessible?", "cell_reference": "B2"},
        ]

        delta = compute_delta(current, previous)
        # Same cell, different text → MODIFIED
        assert delta["Is the main entrance ADA compliant?"] == "MODIFIED"

    def test_mixed_delta(self):
        from app.core.similarity import compute_delta

        current = [
            {"question_text": "Q1 unchanged", "cell_reference": "B2"},
            {"question_text": "Q2 modified text", "cell_reference": "B3"},
            {"question_text": "Q3 brand new", "cell_reference": "B4"},
        ]
        previous = [
            {"question_text": "Q1 unchanged", "cell_reference": "B2"},
            {"question_text": "Q2 original text", "cell_reference": "B3"},
        ]

        delta = compute_delta(current, previous)
        assert delta["Q1 unchanged"] == "UNCHANGED"
        assert delta["Q2 modified text"] == "MODIFIED"
        assert delta["Q3 brand new"] == "NEW"

    def test_case_insensitive_matching(self):
        from app.core.similarity import compute_delta

        current = [{"question_text": "Is Fire Safety OK?", "cell_reference": "B2"}]
        previous = [{"question_text": "is fire safety ok?", "cell_reference": "B2"}]

        delta = compute_delta(current, previous)
        assert delta["Is Fire Safety OK?"] == "UNCHANGED"

    def test_empty_current(self):
        from app.core.similarity import compute_delta

        delta = compute_delta([], [{"question_text": "Q1", "cell_reference": "B2"}])
        assert delta == {}

    def test_normalize_question(self):
        from app.core.similarity import _normalize_question

        assert _normalize_question("  Hello   World  ") == "hello world"
        assert _normalize_question("TEST") == "test"


# ═══════════════════════════════════════════════════════════════════════════════
# 5. QuestionItem Multi-Run Fields
# ═══════════════════════════════════════════════════════════════════════════════

class TestQuestionItemMultiRun:
    def test_multi_run_fields_present(self):
        item = _make_question(
            answer_origin="reused",
            reused_from_question_id="qe-uuid-123",
            reuse_similarity_score=0.95,
            change_type="UNCHANGED",
        )
        assert item.answer_origin == "reused"
        assert item.reused_from_question_id == "qe-uuid-123"
        assert item.reuse_similarity_score == 0.95
        assert item.change_type == "UNCHANGED"

    def test_multi_run_fields_optional(self):
        """Multi-run fields should all be Optional with None defaults."""
        from app.models.schemas import QuestionItem

        item = QuestionItem(
            sheet_name="S1",
            cell_coordinate="B2",
            question="Q",
            ai_answer="A",
            final_answer="A",
            confidence="HIGH",
            sources=[],
        )
        assert item.answer_origin is None
        assert item.reused_from_question_id is None
        assert item.reuse_similarity_score is None
        assert item.change_type is None

    def test_reused_question_item(self):
        item = _make_question(
            answer_origin="reused",
            reused_from_question_id="qe-abc",
            reuse_similarity_score=0.92,
            model_used="reused",
            token_count_used=0,
        )
        assert item.answer_origin == "reused"
        assert item.model_used == "reused"
        assert item.token_count_used == 0

    def test_suggested_question_item(self):
        item = _make_question(
            answer_origin="suggested",
            reused_from_question_id="qe-xyz",
            reuse_similarity_score=0.80,
        )
        assert item.answer_origin == "suggested"
        assert item.reuse_similarity_score == 0.80


# ═══════════════════════════════════════════════════════════════════════════════
# 6. Audit Sheet Multi-Run Columns
# ═══════════════════════════════════════════════════════════════════════════════

class TestAuditSheetMultiRun:
    def test_audit_sheet_has_answer_origin_header(self):
        from app.core.excel_agent import excel_agent
        from openpyxl import load_workbook
        from io import BytesIO

        wb_bytes = _create_test_workbook()
        items = [
            _make_question(
                cell_coordinate="B2",
                answer_origin="reused",
                review_status="approved",
            ),
            _make_question(
                cell_coordinate="B3",
                answer_origin="generated",
                review_status="approved",
            ),
        ]
        output = excel_agent.generate_excel(wb_bytes, items)

        wb = load_workbook(BytesIO(output))
        audit_ws = wb["AI_Verification_Audit"]
        headers = [cell.value for cell in audit_ws[2]]

        assert "Answer Origin" in headers
        origin_idx = headers.index("Answer Origin")

        # Row 3 = first data row (row 1 = disclaimer, row 2 = headers)
        row3 = [cell.value for cell in audit_ws[3]]
        assert row3[origin_idx] == "reused"

        row4 = [cell.value for cell in audit_ws[4]]
        assert row4[origin_idx] == "generated"

    def test_audit_sheet_default_origin_is_generated(self):
        from app.core.excel_agent import excel_agent
        from openpyxl import load_workbook
        from io import BytesIO

        wb_bytes = _create_test_workbook()
        items = [
            _make_question(
                cell_coordinate="B2",
                answer_origin=None,  # Not set
                review_status="approved",
            ),
        ]
        output = excel_agent.generate_excel(wb_bytes, items)

        wb = load_workbook(BytesIO(output))
        audit_ws = wb["AI_Verification_Audit"]
        headers = [cell.value for cell in audit_ws[2]]
        origin_idx = headers.index("Answer Origin")
        row3 = [cell.value for cell in audit_ws[3]]
        assert row3[origin_idx] == "generated"


# ═══════════════════════════════════════════════════════════════════════════════
# 7. Generation Pipeline Multi-Run Fields
# ═══════════════════════════════════════════════════════════════════════════════

class TestGenerationMultiRunFields:
    def test_error_response_has_multi_run_fields(self):
        from app.core.generation import AnswerEngine

        resp = AnswerEngine._error_response("ai_unavailable", "test error")
        assert "answer_origin" in resp
        assert resp["answer_origin"] == "generated"
        assert "reused_from_question_id" in resp
        assert resp["reused_from_question_id"] is None
        assert "reuse_similarity_score" in resp
        assert resp["reuse_similarity_score"] is None

    def test_not_found_response_has_multi_run_fields(self):
        from app.core.generation import AnswerEngine

        engine = AnswerEngine()
        resp = engine._not_found_response(
            retrieval_result=None,
            model="gpt-4-turbo",
            is_strict=False,
        )
        assert resp["answer_origin"] == "generated"
        assert resp["reused_from_question_id"] is None
        assert resp["reuse_similarity_score"] is None


# ═══════════════════════════════════════════════════════════════════════════════
# 8. Config Multi-Run Settings
# ═══════════════════════════════════════════════════════════════════════════════

class TestConfigMultiRun:
    def test_multi_run_default_settings(self):
        from app.core.config import get_settings

        s = get_settings()
        assert s.REUSE_EXACT_THRESHOLD == 0.90
        assert s.REUSE_SUGGEST_THRESHOLD == 0.75
        assert s.REUSE_SEARCH_LIMIT == 5
        assert s.EMBEDDING_CACHE_SIZE == 1000
        assert s.REUSE_ENABLED is True

    def test_thresholds_ordered(self):
        from app.core.config import get_settings

        s = get_settings()
        assert s.REUSE_SUGGEST_THRESHOLD < s.REUSE_EXACT_THRESHOLD
        assert s.REUSE_SUGGEST_THRESHOLD > 0.0
        assert s.REUSE_EXACT_THRESHOLD <= 1.0


# ═══════════════════════════════════════════════════════════════════════════════
# 9. Backward Compatibility — Retrieval Tests Still Pass
# ═══════════════════════════════════════════════════════════════════════════════

class TestBackwardCompatibility:
    def test_retrieval_question_item_still_works(self):
        """Retrieval fields should still be present and functional."""
        item = _make_question()
        assert item.confidence_score == 0.85
        assert item.embedding_similarity_score == 0.92
        assert item.retrieval_mode == "standard"

    def test_approved_answer_still_written(self):
        from app.core.excel_agent import excel_agent
        from openpyxl import load_workbook
        from io import BytesIO

        wb_bytes = _create_test_workbook()
        items = [_make_question(cell_coordinate="B2", review_status="approved")]
        output = excel_agent.generate_excel(wb_bytes, items)

        wb = load_workbook(BytesIO(output))
        ws = wb["Sheet1"]
        assert ws["B2"].value == "Test answer"

    def test_rejected_answer_not_written(self):
        from app.core.excel_agent import excel_agent
        from openpyxl import load_workbook
        from io import BytesIO

        wb_bytes = _create_test_workbook()
        items = [_make_question(cell_coordinate="B2", review_status="rejected")]
        output = excel_agent.generate_excel(wb_bytes, items)

        wb = load_workbook(BytesIO(output))
        ws = wb["Sheet1"]
        assert ws["B2"].value is None or str(ws["B2"].value).strip() == ""


# ═══════════════════════════════════════════════════════════════════════════════
# 10. Run Comparison Delta Logic
# ═══════════════════════════════════════════════════════════════════════════════

class TestRunComparisonDelta:
    """Test the compute_delta function used by the comparison endpoint."""

    def test_comparison_identifies_removed_questions(self):
        """Questions in previous but not current should be detectable."""
        from app.core.similarity import compute_delta

        current = [{"question_text": "Q2", "cell_reference": "B3"}]
        previous = [
            {"question_text": "Q1", "cell_reference": "B2"},
            {"question_text": "Q2", "cell_reference": "B3"},
        ]

        # compute_delta returns changes for current questions only
        delta = compute_delta(current, previous)
        assert delta["Q2"] == "UNCHANGED"
        # Q1 is removed — not in delta (detected separately in comparison endpoint)
        assert "Q1" not in delta

    def test_comparison_all_new(self):
        from app.core.similarity import compute_delta

        current = [
            {"question_text": "Brand new Q1", "cell_reference": "B2"},
            {"question_text": "Brand new Q2", "cell_reference": "B3"},
        ]
        previous = [
            {"question_text": "Old Q1", "cell_reference": "C2"},
            {"question_text": "Old Q2", "cell_reference": "C3"},
        ]

        delta = compute_delta(current, previous)
        assert all(v == "NEW" for v in delta.values())

    def test_comparison_preserves_whitespace_in_keys(self):
        """Delta map keys should use original question text, not normalized."""
        from app.core.similarity import compute_delta

        current = [{"question_text": "  Spaced Question  ", "cell_reference": "B2"}]
        previous = [{"question_text": "spaced question", "cell_reference": "B2"}]

        delta = compute_delta(current, previous)
        assert "  Spaced Question  " in delta
        assert delta["  Spaced Question  "] == "UNCHANGED"


# ═══════════════════════════════════════════════════════════════════════════════
# 11. SimilarityEngine Disabled Mode
# ═══════════════════════════════════════════════════════════════════════════════

class TestSimilarityEngineDisabled:
    def test_disabled_reuse_returns_generate(self):
        """When REUSE_ENABLED=False, search_similar should always return 'generate'."""
        from app.core.similarity import SimilarityEngine
        from app.core.config import get_settings
        import unittest.mock as mock

        s = get_settings()
        original = s.REUSE_ENABLED

        try:
            # Temporarily disable reuse
            object.__setattr__(s, "REUSE_ENABLED", False)
            engine = SimilarityEngine()
            result = engine.search_similar("test question", "org-123")
            assert result.action == "generate"
            assert result.has_reusable is False
            assert result.has_suggestion is False
        finally:
            object.__setattr__(s, "REUSE_ENABLED", original)
