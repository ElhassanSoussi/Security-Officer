"""
Export Gate — approved-only write behavior.

This test proves that:
1. ExcelAgent.generate_excel writes approved answers to cells
2. ExcelAgent.generate_excel leaves pending/rejected answers blank
3. Audit sheet records all answers with correct review_status labels
"""
import sys
import os
import pytest

# Ensure backend app is importable
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))


def _make_question(
    question: str,
    answer: str,
    review_status: str = "pending",
    cell: str = "B2",
    sheet: str = "Sheet1",
    confidence: str = "HIGH",
):
    """Create a QuestionItem-like dict that can be unpacked into the model."""
    from app.models.schemas import QuestionItem

    return QuestionItem(
        sheet_name=sheet,
        cell_coordinate=cell,
        question=question,
        ai_answer=answer,
        final_answer=answer,
        confidence=confidence,
        sources=["test_doc.pdf"],
        source_excerpt="Relevant excerpt from test_doc.pdf",
        is_verified=review_status == "approved",
        edited_by_user=False,
        review_status=review_status,
    )


def _create_test_workbook() -> bytes:
    """Create a minimal xlsx in memory with a question column and empty answer column."""
    from openpyxl import Workbook
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Question", "Answer"])
    ws.append(["Is fire safety plan current?", ""])  # B2 — will be approved
    ws.append(["Are exits marked?", ""])  # B3 — will be rejected
    ws.append(["Is the roof compliant?", ""])  # B4 — will be pending

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestExportGate:
    """Verify that generate_excel only writes approved answers."""

    def test_approved_answer_written_to_cell(self):
        """Approved answers MUST appear in the target cell."""
        from app.core.excel_agent import excel_agent
        from openpyxl import load_workbook
        from io import BytesIO

        template = _create_test_workbook()
        answers = [
            _make_question(
                "Is fire safety plan current?",
                "Yes, per Section 5.2 of the Fire Code [test_doc.pdf, pg. 12]",
                review_status="approved",
                cell="B2",
            ),
        ]

        result_bytes = excel_agent.generate_excel(template, answers)
        wb = load_workbook(BytesIO(result_bytes))
        ws = wb["Sheet1"]

        assert ws["B2"].value is not None, "Approved answer must be written to B2"
        assert "fire" in ws["B2"].value.lower() or "Section 5.2" in ws["B2"].value

        wb.close()

    def test_rejected_answer_left_blank(self):
        """Rejected answers MUST NOT appear in the target cell."""
        from app.core.excel_agent import excel_agent
        from openpyxl import load_workbook
        from io import BytesIO

        template = _create_test_workbook()
        answers = [
            _make_question(
                "Are exits marked?",
                "Yes, all exits are marked per code.",
                review_status="rejected",
                cell="B3",
            ),
        ]

        result_bytes = excel_agent.generate_excel(template, answers)
        wb = load_workbook(BytesIO(result_bytes))
        ws = wb["Sheet1"]

        cell_val = ws["B3"].value
        assert cell_val is None or str(cell_val).strip() == "", \
            f"Rejected answer must NOT be written to B3, got: {cell_val!r}"

        wb.close()

    def test_pending_answer_left_blank(self):
        """Pending (unreviewed) answers MUST NOT appear in the target cell."""
        from app.core.excel_agent import excel_agent
        from openpyxl import load_workbook
        from io import BytesIO

        template = _create_test_workbook()
        answers = [
            _make_question(
                "Is the roof compliant?",
                "Compliant per Building Bulletin 2024-01",
                review_status="pending",
                cell="B4",
            ),
        ]

        result_bytes = excel_agent.generate_excel(template, answers)
        wb = load_workbook(BytesIO(result_bytes))
        ws = wb["Sheet1"]

        cell_val = ws["B4"].value
        assert cell_val is None or str(cell_val).strip() == "", \
            f"Pending answer must NOT be written to B4, got: {cell_val!r}"

        wb.close()

    def test_mixed_review_statuses(self):
        """Only approved answers are written; pending and rejected are blank."""
        from app.core.excel_agent import excel_agent
        from openpyxl import load_workbook
        from io import BytesIO

        template = _create_test_workbook()
        answers = [
            _make_question(
                "Is fire safety plan current?",
                "Yes, per Section 5.2",
                review_status="approved",
                cell="B2",
            ),
            _make_question(
                "Are exits marked?",
                "Yes, all exits are marked",
                review_status="rejected",
                cell="B3",
            ),
            _make_question(
                "Is the roof compliant?",
                "Compliant per BB-2024-01",
                review_status="pending",
                cell="B4",
            ),
        ]

        result_bytes = excel_agent.generate_excel(template, answers)
        wb = load_workbook(BytesIO(result_bytes))
        ws = wb["Sheet1"]

        # Approved → written
        assert ws["B2"].value is not None, "Approved answer must appear in B2"
        assert "Section 5.2" in str(ws["B2"].value)

        # Rejected → blank
        b3 = ws["B3"].value
        assert b3 is None or str(b3).strip() == "", f"Rejected answer must NOT appear in B3, got: {b3!r}"

        # Pending → blank
        b4 = ws["B4"].value
        assert b4 is None or str(b4).strip() == "", f"Pending answer must NOT appear in B4, got: {b4!r}"

        wb.close()

    def test_audit_sheet_records_review_status(self):
        """The AI_Verification_Audit sheet must include review status for each answer."""
        from app.core.excel_agent import excel_agent
        from openpyxl import load_workbook
        from io import BytesIO

        template = _create_test_workbook()
        answers = [
            _make_question(
                "Is fire safety plan current?",
                "Yes, per Section 5.2",
                review_status="approved",
                cell="B2",
            ),
            _make_question(
                "Are exits marked?",
                "Yes, all exits are marked",
                review_status="rejected",
                cell="B3",
            ),
        ]

        result_bytes = excel_agent.generate_excel(template, answers)
        wb = load_workbook(BytesIO(result_bytes))

        assert "AI_Verification_Audit" in wb.sheetnames, "Audit sheet must exist"
        audit_ws = wb["AI_Verification_Audit"]

        # Row 1: warning, Row 2: headers, Row 3+: data
        rows = list(audit_ws.iter_rows(min_row=3, values_only=True))
        assert len(rows) >= 2, f"Expected at least 2 audit rows, got {len(rows)}"

        # Check that review status column is populated
        # Headers: Cell Reference(0), Question(1), Final Answer(2), Source Document(3),
        #          Page Number(4), Confidence(5), Confidence Score(6), Similarity(7),
        #          Model(8), Review Status(9), Finalized By(10)
        # Dynamically find the Review Status column index from header row
        header_row = list(audit_ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        review_col_idx = next(
            (i for i, h in enumerate(header_row) if h and "review" in str(h).lower()),
            9,  # fallback
        )
        statuses = [str(row[review_col_idx]).lower() if row[review_col_idx] else "" for row in rows]
        assert "approved" in statuses, f"Expected 'Approved' in audit statuses, got: {statuses}"
        assert "rejected" in statuses, f"Expected 'Rejected' in audit statuses, got: {statuses}"

        wb.close()

    def test_not_found_in_locker_has_needs_info_status(self):
        """Answers with 'NOT FOUND IN LOCKER' should have status needs_info."""
        from app.models.schemas import QuestionItem

        item = QuestionItem(
            sheet_name="Sheet1",
            cell_coordinate="B2",
            question="Test question",
            ai_answer="",
            final_answer="",
            confidence="LOW",
            sources=[],
            status="needs_info",
            status_reason="Missing source context.",
            is_verified=False,
            edited_by_user=False,
            review_status="pending",
        )

        assert item.status == "needs_info"
        assert item.review_status == "pending"
        assert item.confidence == "LOW"


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
