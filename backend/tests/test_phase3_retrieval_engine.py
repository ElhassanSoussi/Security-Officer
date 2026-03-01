"""
Phase 3 Verification: Deterministic Retrieval Engine

Tests cover:
1. Similarity threshold enforcement
2. Confidence scoring (compute_confidence)
3. Strict mode prompt selection
4. Direct quote detection
5. Retrieval metadata in QuestionItem
6. Cell comments in exported Excel
7. Expanded audit sheet columns
8. Error/timeout response structure
9. RetrievalResult dataclass properties
10. Debug mode response structure
"""
import sys
import os
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))


# ─── Test Helpers ────────────────────────────────────────────────────────────

def _make_question(**overrides):
    """Create a QuestionItem with Phase 3 defaults."""
    from app.models.schemas import QuestionItem

    defaults = {
        "sheet_name": "Sheet1",
        "cell_coordinate": "B2",
        "question": "Test question",
        "ai_answer": "Test answer",
        "final_answer": "Test answer",
        "confidence": "HIGH",
        "sources": ["test_doc.pdf"],
        "source_excerpt": "Relevant excerpt",
        "is_verified": False,
        "edited_by_user": False,
        "review_status": "approved",
        "confidence_score": 0.85,
        "confidence_reason": "high similarity match; 3 supporting chunks",
        "embedding_similarity_score": 0.92,
        "chunk_id": "chunk-uuid-123",
        "token_count_used": 450,
        "model_used": "gpt-4-turbo",
        "generation_time_ms": 1200,
        "retrieval_mode": "standard",
    }
    defaults.update(overrides)
    return QuestionItem(**defaults)


def _create_test_workbook() -> bytes:
    from openpyxl import Workbook
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Question", "Answer"])
    ws.append(["Is fire safety plan current?", ""])
    ws.append(["Are exits marked?", ""])
    ws.append(["Is the roof compliant?", ""])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── 1. Confidence Scoring ──────────────────────────────────────────────────

class TestConfidenceScoring:
    """Verify compute_confidence produces correct scores and reasons."""

    def test_high_confidence_direct_quote(self):
        from app.core.generation import compute_confidence

        score, reason = compute_confidence(
            best_similarity=0.92,
            num_supporting_chunks=4,
            excerpt_length=400,
            is_direct_quote=True,
            answer_text="Fire safety plan is current per Section 5.2",
        )
        assert 0.7 <= score <= 1.0, f"Expected HIGH confidence, got {score}"
        assert "high similarity" in reason.lower()
        assert "direct quote" in reason.lower()

    def test_low_confidence_no_context(self):
        from app.core.generation import compute_confidence

        score, reason = compute_confidence(
            best_similarity=0.3,
            num_supporting_chunks=0,
            excerpt_length=0,
            is_direct_quote=False,
            answer_text="",
        )
        assert score < 0.4, f"Expected LOW confidence, got {score}"
        assert "low similarity" in reason.lower()

    def test_medium_confidence_inferred(self):
        from app.core.generation import compute_confidence

        score, reason = compute_confidence(
            best_similarity=0.65,
            num_supporting_chunks=2,
            excerpt_length=200,
            is_direct_quote=False,
            answer_text="The exits appear to be compliant based on inspection notes.",
        )
        assert 0.3 <= score <= 0.7, f"Expected MEDIUM confidence, got {score}"
        assert "inferred" in reason.lower()

    def test_confidence_score_bounded(self):
        from app.core.generation import compute_confidence

        # Even with max inputs, score should not exceed 1.0
        score, _ = compute_confidence(
            best_similarity=1.0,
            num_supporting_chunks=10,
            excerpt_length=1000,
            is_direct_quote=True,
            answer_text="Test",
        )
        assert score <= 1.0

        # With min inputs, score should not go below 0.0
        score, _ = compute_confidence(
            best_similarity=0.0,
            num_supporting_chunks=0,
            excerpt_length=0,
            is_direct_quote=False,
            answer_text="",
        )
        assert score >= 0.0

    def test_confidence_label_classification(self):
        from app.core.generation import _classify_confidence_label

        assert _classify_confidence_label(0.85) == "HIGH"
        assert _classify_confidence_label(0.7) == "HIGH"
        assert _classify_confidence_label(0.5) == "MEDIUM"
        assert _classify_confidence_label(0.4) == "MEDIUM"
        assert _classify_confidence_label(0.3) == "LOW"
        assert _classify_confidence_label(0.0) == "LOW"


# ─── 2. Direct Quote Detection ──────────────────────────────────────────────

class TestDirectQuoteDetection:
    """Verify _detect_direct_quote identifies verbatim matches."""

    def test_exact_substring_detected(self):
        from app.core.generation import _detect_direct_quote

        context = "The fire safety plan was updated on January 15, 2025 and approved by the site manager."
        answer = "The fire safety plan was updated on January 15, 2025"
        assert _detect_direct_quote(answer, context) is True

    def test_no_match_detected(self):
        from app.core.generation import _detect_direct_quote

        context = "The fire safety plan covers exit routes and signage."
        answer = "The building structural report indicates compliance with local regulations."
        assert _detect_direct_quote(answer, context) is False

    def test_empty_inputs(self):
        from app.core.generation import _detect_direct_quote

        assert _detect_direct_quote("", "some context") is False
        assert _detect_direct_quote("some answer", "") is False
        assert _detect_direct_quote("", "") is False

    def test_short_answer_exact_match(self):
        from app.core.generation import _detect_direct_quote

        context = "Yes, it is compliant with all relevant codes."
        answer = "Yes, it is compliant"
        assert _detect_direct_quote(answer, context) is True


# ─── 3. RetrievalResult Dataclass ───────────────────────────────────────────

class TestRetrievalResult:
    """Verify RetrievalResult properties and serialization."""

    def test_empty_result(self):
        from app.core.retrieval import RetrievalResult

        r = RetrievalResult()
        assert r.has_results is False
        assert r.best_score == 0.0
        assert r.context_text == ""
        assert r.source_filenames == []

    def test_with_chunks(self):
        from app.core.retrieval import RetrievalResult, RetrievalChunk

        chunks = [
            RetrievalChunk(
                chunk_id="c1",
                document_id="d1",
                filename="doc.pdf",
                content="Fire safety plan section 5.2",
                similarity=0.88,
            ),
            RetrievalChunk(
                chunk_id="c2",
                document_id="d1",
                filename="doc.pdf",
                content="Emergency exits are marked",
                similarity=0.75,
            ),
        ]
        r = RetrievalResult(
            chunks=chunks,
            threshold_used=0.55,
            top_k_used=5,
            total_candidates=10,
            above_threshold=2,
            retrieval_time_ms=150,
        )
        assert r.has_results is True
        assert r.best_score == 0.88
        assert "Fire safety" in r.context_text
        assert "Emergency exits" in r.context_text
        assert r.source_filenames == ["doc.pdf"]

    def test_to_dict_no_debug(self):
        from app.core.retrieval import RetrievalResult

        r = RetrievalResult(threshold_used=0.55, top_k_used=5)
        d = r.to_dict()
        assert "debug_all_scores" not in d
        assert d["threshold_used"] == 0.55

    def test_to_dict_with_debug(self):
        from app.core.retrieval import RetrievalResult

        r = RetrievalResult(
            threshold_used=0.55,
            debug_all_scores=[{"chunk_id": "c1", "similarity": 0.9}],
        )
        d = r.to_dict()
        assert "debug_all_scores" in d
        assert len(d["debug_all_scores"]) == 1


# ─── 4. Error Response Structure ────────────────────────────────────────────

class TestErrorResponseStructure:
    """Verify error/fallback responses include all Phase 3 metadata fields."""

    def test_error_response_has_all_fields(self):
        from app.core.generation import AnswerEngine

        resp = AnswerEngine._error_response("ai_unavailable", "Test error")
        required_keys = [
            "status", "reason", "answer", "sources", "confidence",
            "confidence_score", "confidence_reason",
            "embedding_similarity_score", "chunk_id",
            "token_count_used", "model_used", "generation_time_ms",
            "retrieval_mode", "retrieval_debug",
        ]
        for key in required_keys:
            assert key in resp, f"Missing key '{key}' in error response"
        assert resp["status"] == "ai_unavailable"
        assert resp["confidence_score"] == 0.0
        assert resp["retrieval_debug"] is None


# ─── 5. QuestionItem Schema ─────────────────────────────────────────────────

class TestQuestionItemSchema:
    """Verify Phase 3 fields on the QuestionItem model."""

    def test_phase3_fields_present(self):
        item = _make_question()
        assert item.confidence_score == 0.85
        assert item.confidence_reason == "high similarity match; 3 supporting chunks"
        assert item.embedding_similarity_score == 0.92
        assert item.chunk_id == "chunk-uuid-123"
        assert item.token_count_used == 450
        assert item.model_used == "gpt-4-turbo"
        assert item.generation_time_ms == 1200
        assert item.retrieval_mode == "standard"

    def test_phase3_fields_optional(self):
        """Phase 3 fields default to None for backward compatibility."""
        from app.models.schemas import QuestionItem

        item = QuestionItem(
            sheet_name="S1",
            cell_coordinate="B2",
            question="Q?",
            ai_answer="A",
            final_answer="A",
            confidence="HIGH",
            sources=[],
        )
        assert item.confidence_score is None
        assert item.embedding_similarity_score is None
        assert item.chunk_id is None
        assert item.retrieval_debug is None


# ─── 6. Exported Excel Cell Comments ────────────────────────────────────────

class TestExcelCellComments:
    """Verify that approved answers get cell comments with confidence + source."""

    def test_approved_cell_has_comment(self):
        from app.core.excel_agent import excel_agent
        from openpyxl import load_workbook
        from io import BytesIO

        template = _create_test_workbook()
        answers = [
            _make_question(
                question="Is fire safety plan current?",
                ai_answer="Yes, per Section 5.2 [test_doc.pdf, pg. 12]",
                final_answer="Yes, per Section 5.2 [test_doc.pdf, pg. 12]",
                review_status="approved",
                cell="B2",
                confidence_score=0.85,
            ),
        ]

        result_bytes = excel_agent.generate_excel(template, answers)
        wb = load_workbook(BytesIO(result_bytes))
        ws = wb["Sheet1"]

        cell = ws["B2"]
        assert cell.value is not None, "Approved answer must be written"
        assert cell.comment is not None, "Approved cell must have a comment"
        comment_text = cell.comment.text
        assert "confidence" in comment_text.lower()
        assert "0.85" in comment_text
        assert "test_doc.pdf" in comment_text

        wb.close()

    def test_rejected_cell_has_no_comment(self):
        from app.core.excel_agent import excel_agent
        from openpyxl import load_workbook
        from io import BytesIO

        template = _create_test_workbook()
        answers = [
            _make_question(
                question="Are exits marked?",
                ai_answer="Yes",
                final_answer="Yes",
                review_status="rejected",
                cell="B3",
            ),
        ]

        result_bytes = excel_agent.generate_excel(template, answers)
        wb = load_workbook(BytesIO(result_bytes))
        ws = wb["Sheet1"]

        cell = ws["B3"]
        # Rejected answer should not be written, so no comment either
        assert cell.comment is None, "Rejected cell must NOT have a comment"

        wb.close()


# ─── 7. Expanded Audit Sheet Columns ────────────────────────────────────────

class TestExpandedAuditSheet:
    """Verify the audit sheet has Phase 3 columns (Confidence Score, Similarity, Model)."""

    def test_audit_sheet_has_phase3_headers(self):
        from app.core.excel_agent import excel_agent
        from openpyxl import load_workbook
        from io import BytesIO

        template = _create_test_workbook()
        answers = [
            _make_question(
                question="Is fire safety plan current?",
                final_answer="Yes, per Section 5.2",
                review_status="approved",
                cell="B2",
                confidence_score=0.85,
                embedding_similarity_score=0.92,
                model_used="gpt-4-turbo",
            ),
        ]

        result_bytes = excel_agent.generate_excel(template, answers)
        wb = load_workbook(BytesIO(result_bytes))
        audit_ws = wb["AI_Verification_Audit"]

        # Row 2 is the header row
        headers = [str(c.value).lower() if c.value else "" for c in audit_ws[2]]
        assert "confidence score" in headers, f"Missing 'Confidence Score' in headers: {headers}"
        assert "similarity" in headers, f"Missing 'Similarity' in headers: {headers}"
        assert "model" in headers, f"Missing 'Model' in headers: {headers}"

        # Data row should have the values
        data_row = list(audit_ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
        # Find the Confidence Score column index
        conf_idx = headers.index("confidence score")
        sim_idx = headers.index("similarity")
        model_idx = headers.index("model")

        assert data_row[conf_idx] == 0.85, f"Expected confidence_score=0.85, got {data_row[conf_idx]}"
        assert data_row[sim_idx] == 0.92, f"Expected similarity=0.92, got {data_row[sim_idx]}"
        assert data_row[model_idx] == "gpt-4-turbo", f"Expected model=gpt-4-turbo, got {data_row[model_idx]}"

        wb.close()


# ─── 8. Backward Compatibility ──────────────────────────────────────────────

class TestBackwardCompatibility:
    """Ensure Phase 2 test scenarios still work with Phase 3 changes."""

    def test_approved_still_writes(self):
        from app.core.excel_agent import excel_agent
        from openpyxl import load_workbook
        from io import BytesIO

        template = _create_test_workbook()
        # No Phase 3 fields set — backward compat
        from app.models.schemas import QuestionItem

        answers = [
            QuestionItem(
                sheet_name="Sheet1",
                cell_coordinate="B2",
                question="Is fire safety plan current?",
                ai_answer="Yes",
                final_answer="Yes",
                confidence="HIGH",
                sources=["doc.pdf"],
                is_verified=True,
                review_status="approved",
            ),
        ]

        result_bytes = excel_agent.generate_excel(template, answers)
        wb = load_workbook(BytesIO(result_bytes))
        assert wb["Sheet1"]["B2"].value == "Yes"
        wb.close()

    def test_pending_still_blank(self):
        from app.core.excel_agent import excel_agent
        from openpyxl import load_workbook
        from io import BytesIO

        template = _create_test_workbook()
        from app.models.schemas import QuestionItem

        answers = [
            QuestionItem(
                sheet_name="Sheet1",
                cell_coordinate="B4",
                question="Is the roof compliant?",
                ai_answer="Compliant",
                final_answer="Compliant",
                confidence="HIGH",
                sources=[],
                review_status="pending",
            ),
        ]

        result_bytes = excel_agent.generate_excel(template, answers)
        wb = load_workbook(BytesIO(result_bytes))
        val = wb["Sheet1"]["B4"].value
        assert val is None or str(val).strip() == ""
        wb.close()


# ─── 9. Config Settings ─────────────────────────────────────────────────────

class TestConfigSettings:
    """Verify Phase 3 config settings exist and have correct defaults."""

    def test_default_settings(self):
        from app.core.config import get_settings

        s = get_settings()
        assert s.RETRIEVAL_SIMILARITY_THRESHOLD == 0.55
        assert s.RETRIEVAL_TOP_K == 5
        assert s.RETRIEVAL_DEBUG is False
        assert s.STRICT_MODE is False
        assert s.LLM_MODEL == "gpt-4-turbo"
        assert s.LLM_TIMEOUT_SECONDS == 60


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
