from io import BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
from typing import List, Optional, Dict, Any
from app.models.schemas import QuestionItem
from app.core.generation import answer_engine
from app.core.config import get_settings
try:
    from openai import OpenAI
except ImportError:
    class OpenAI:
        def __init__(self, api_key=None): self.chat = type("obj", (), {"completions": type("obj", (), {"create": lambda **k: type("obj", (), {"choices": [type("obj", (), {"message": type("obj", (), {"content": "Mock excel", "tool_calls": []})})]})})})
import re
import json

settings = get_settings()

class ExcelAgent:
    def __init__(self):
        self.client = OpenAI(api_key=settings.OPENAI_API_KEY)

    def _load_workbook(self, file_content: bytes):
        """Return an open workbook and the backing buffer.

        Keeping the BytesIO stream alive for the lifetime of the workbook
        prevents `ValueError: I/O operation on closed file` that can surface
        when openpyxl lazily reads parts of the file (common with large/merged
        worksheets).
        """
        buffer = BytesIO(file_content)
        wb = load_workbook(filename=buffer, data_only=False, keep_vba=True)
        return wb, buffer

    def _sheet_snapshot(self, ws, max_rows: int = 20, max_cols: int = 15) -> List[List[Dict[str, str]]]:
        """Lightweight representation of the top of the sheet for LLM layout detection.
        Skips hidden rows/columns to avoid confusing the layout detector."""
        snapshot: List[List[Dict[str, str]]] = []
        rows_seen = 0
        for row in ws.iter_rows(min_row=1, max_col=max_cols):
            if rows_seen >= max_rows:
                break
            row_num = row[0].row if row else None
            if row_num is None:
                continue
            # Skip hidden rows
            if ws.row_dimensions[row_num].hidden:
                continue
            row_repr = []
            for cell in row:
                # Skip hidden columns
                col_letter = cell.column_letter
                if ws.column_dimensions[col_letter].hidden:
                    continue
                val = "" if cell.value is None else str(cell.value)
                row_repr.append({
                    "cell": f"{col_letter}{cell.row}",
                    "value": val
                })
            if row_repr:
                snapshot.append(row_repr)
                rows_seen += 1
        return snapshot

    def analyze_sheet_structure(self, rows_snapshot: List[List[Dict[str, str]]], merged_ranges: List[str]) -> Dict[str, Any]:
        """
        Ask the LLM to infer layout (question/answer columns and start row) using
        a coordinate-aware snapshot plus merged-cell hints. This avoids brittle
        fixed-column heuristics on city forms with merged headers.
        """
        prompt = f"""
        You are inspecting the first ~20 rows of an Excel safety questionnaire. Each cell is shown
        with its coordinate. Merged ranges are provided. Identify which column holds the question
        text and which column should store the answers. Return zero-based column indexes.

        Snapshot:
        {json.dumps(rows_snapshot)}

        Merged Ranges (A1:B1 style): {merged_ranges}

        Return STRICT JSON:
        {{
            "question_col_index": <int>,
            "answer_col_index": <int>,
            "start_row_index": <int>,  // zero-based data row where questions begin
            "confidence": "HIGH" | "LOW"
        }}
        Rules:
        - Prefer columns with headers like Question/Requirement for questions and Answer/Response for answers.
        - If answers appear in the same column (self-filled), set question_col_index to that column and answer_col_index to the next likely blank column.
        - For vertical layouts where answers are directly below the question, set answer_col_index equal to question_col_index and start_row_index accordingly.
        """

        try:
            response = self.client.chat.completions.create(
                model="gpt-4-turbo",
                messages=[
                    {"role": "system", "content": "You are a Data Analyst specialized in parsing security questionnaires."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0,
                response_format={"type": "json_object"}
            )
            return json.loads(response.choices[0].message.content)
        except Exception as e:
            print(f"⚠️ Sheet Analysis Failed: {e}")
            # Fallback to conservative defaults if LLM fails
            return {"question_col_index": 0, "answer_col_index": 1, "start_row_index": 1, "confidence": "LOW"}

    def analyze_excel(
        self,
        file_content: bytes,
        org_id: str,
        project_id: Optional[str],
        token: Optional[str] = None,
    ) -> List[QuestionItem]:
        results: List[QuestionItem] = []

        wb, buffer = self._load_workbook(file_content)
        try:
            for sheet_name in wb.sheetnames:
                if sheet_name == "AI_Verification_Audit":
                    continue

                ws = wb[sheet_name]

                # Skip hidden sheets
                if ws.sheet_state != "visible":
                    print(f"⏭️ Skipping hidden sheet: {sheet_name}")
                    continue

                # Guard against empty/corrupt sheets
                sheet_max_row = ws.max_row
                sheet_max_col = ws.max_column
                if not sheet_max_row or sheet_max_row < 2:
                    print(f"⏭️ Skipping empty/tiny sheet: {sheet_name}")
                    continue
                if not sheet_max_col or sheet_max_col < 1:
                    print(f"⏭️ Skipping sheet with no columns: {sheet_name}")
                    continue

                # 1) Snapshot + merged hints for LLM
                try:
                    snapshot = self._sheet_snapshot(ws)
                    merged_ranges = [str(rng.coord) for rng in ws.merged_cells.ranges][:25]
                except Exception as e:
                    print(f"⚠️ Snapshot failed on sheet {sheet_name}: {e}")
                    continue

                if not snapshot:
                    continue

                # 2) Layout detection
                structure = self.analyze_sheet_structure(snapshot, merged_ranges)

                # Guard: LLM may return malformed JSON or a non-dict.
                if not isinstance(structure, dict):
                    print(f"⚠️ Sheet {sheet_name}: layout detection returned non-dict, using defaults")
                    structure = {"question_col_index": 0, "answer_col_index": 1, "start_row_index": 1, "confidence": "LOW"}

                q_col_idx = structure.get("question_col_index", 0)
                a_col_idx = structure.get("answer_col_index", q_col_idx + 1)
                start_row = structure.get("start_row_index", 1)

                # Normalize + clamp indexes
                def to_int(val, default):
                    try:
                        return max(0, int(val))
                    except Exception:
                        return default

                q_col_idx = to_int(q_col_idx, 0)
                a_col_idx = to_int(a_col_idx, q_col_idx + 1)
                start_row = to_int(start_row, 1)

                # Clamp indexes against actual sheet dimensions to prevent out-of-bounds
                actual_max_col = sheet_max_col or 1
                if q_col_idx >= actual_max_col:
                    q_col_idx = 0
                if a_col_idx >= actual_max_col:
                    a_col_idx = min(q_col_idx + 1, actual_max_col - 1)
                if start_row >= (sheet_max_row or 1):
                    start_row = 1

                # Map merged cells to their master cell for safe value access
                merged_lookup: Dict[str, Any] = {}
                for rng in ws.merged_cells.ranges:
                    master = ws.cell(rng.min_row, rng.min_col)
                    for r in range(rng.min_row, rng.max_row + 1):
                        for c in range(rng.min_col, rng.max_col + 1):
                            merged_lookup[f"{ws.cell(r, c).coordinate}"] = master

                def safe_get(row_cells, idx):
                    return row_cells[idx] if 0 <= idx < len(row_cells) else None

                def effective_value(cell):
                    if cell is None:
                        return None
                    if cell.value is not None:
                        return cell.value
                    master = merged_lookup.get(cell.coordinate)
                    return master.value if master else None

                # openpyxl is 1-based; convert start_row (0-based) -> excel row number
                start_excel_row = start_row + 1
                max_col = max(ws.max_column or 1, a_col_idx + 1)

                for row_offset, row in enumerate(ws.iter_rows(min_row=start_excel_row, max_col=max_col)):
                    # Safety: skip empty/ragged rows that iter_rows can yield
                    if not row:
                        continue
                    sheet_row_num = start_excel_row + row_offset

                    # Skip hidden rows during iteration
                    if ws.row_dimensions[sheet_row_num].hidden:
                        continue

                    try:
                        q_cell = safe_get(row, q_col_idx)
                        a_cell = safe_get(row, a_col_idx)

                        # If answer column is outside current row bounds (ragged row), create a fresh cell
                        if a_cell is None:
                            a_cell = ws.cell(row=sheet_row_num, column=a_col_idx + 1)

                        # Resolve values, accounting for merged cells
                        question_raw = effective_value(q_cell)
                        if question_raw is None:
                            continue

                        text = str(question_raw).strip()
                        if not text or len(text) < 3:
                            continue

                        # Skip if answer already exists
                        ans_val = effective_value(a_cell)
                        if ans_val and str(ans_val).strip():
                            continue

                        print(f"🤖 [Sheet:{sheet_name}] Answering: {text[:60]}...")
                        result = answer_engine.generate_answer(text, org_id, project_id, token=token)

                        item = QuestionItem(
                            sheet_name=sheet_name,
                            cell_coordinate=a_cell.coordinate,
                            question=text,
                            ai_answer=result["answer"],
                            final_answer=result["answer"],
                            confidence=result["confidence"],
                            sources=result["sources"],
                            source_id=result.get("source_id"),
                            source_page=result.get("source_page"),
                            source_excerpt=result.get("source_excerpt"),
                            status=result.get("status"),
                            status_reason=result.get("reason"),
                            is_verified=False,
                            edited_by_user=False,
                            # Phase 3: Structured confidence
                            confidence_score=result.get("confidence_score"),
                            confidence_reason=result.get("confidence_reason"),
                            # Phase 3: Retrieval metadata
                            embedding_similarity_score=result.get("embedding_similarity_score"),
                            chunk_id=result.get("chunk_id"),
                            token_count_used=result.get("token_count_used"),
                            model_used=result.get("model_used"),
                            generation_time_ms=result.get("generation_time_ms"),
                            retrieval_mode=result.get("retrieval_mode"),
                            retrieval_debug=result.get("retrieval_debug"),
                            # Phase 4: Answer reuse metadata
                            answer_origin=result.get("answer_origin"),
                            reused_from_question_id=result.get("reused_from_question_id"),
                            reuse_similarity_score=result.get("reuse_similarity_score"),
                        )
                        results.append(item)
                    except Exception as row_err:
                        print(f"⚠️ Row Error at {sheet_name}:{sheet_row_num} => {row_err}")
                        continue
        except Exception as e:
            print(f"🛑 CRITICAL EXCEL AGENT CRASH: {e}")
            raise e
        finally:
            wb.close()
            buffer.close()

        return results

    def generate_excel(self, file_content: bytes, answers: List[QuestionItem]) -> bytes:
        try:
            wb, in_buffer = self._load_workbook(file_content)
            try:
                # 1. Reset Audit Sheet
                if "AI_Verification_Audit" in wb.sheetnames:
                    del wb["AI_Verification_Audit"]

                audit_ws = wb.create_sheet("AI_Verification_Audit")
                audit_ws.append(["⚠️ AI-GENERATED CONTENT. MANUAL REVIEW REQUIRED BEFORE SUBMISSION."])
                headers = [
                    "Cell Reference", "Question", "Final Answer", "Source Document",
                    "Page Number", "Confidence", "Confidence Score", "Similarity",
                    "Model", "Review Status", "Finalized By", "Answer Origin",
                ]
                audit_ws.append(headers)

                # 2. Apply Answers — only approved answers are written; rejected/pending are left blank.
                for item in answers:
                    review = (item.review_status or "").strip().lower()
                    # Determine if this answer should be written to the cell
                    should_write = review == "approved" or (
                        review not in ("rejected",) and item.is_verified
                    )
                    
                    if item.sheet_name in wb.sheetnames:
                        ws = wb[item.sheet_name]
                        try:
                            cell = ws[item.cell_coordinate]
                            if should_write:
                                # Only write into empty cells to avoid stomping user data
                                if not cell.value or not str(cell.value).strip():
                                    ws[item.cell_coordinate] = item.final_answer
                                    # Phase 3: Add cell comment with confidence + source info
                                    try:
                                        conf_str = f"{item.confidence_score:.2f}" if item.confidence_score is not None else item.confidence
                                        src_name = item.sources[0] if item.sources else "N/A"
                                        comment_text = f"AI generated — confidence: {conf_str} — source: {src_name}"
                                        cell.comment = Comment(comment_text, "NYC Compliance Architect")
                                    except Exception:
                                        pass  # Never let comment failure block export
                                else:
                                    print(f"⚠️ Skipping write to {item.cell_coordinate} - Cell already has data.")
                            else:
                                # Not approved — leave cell blank (do not write)
                                print(f"⏭️ Skipping {item.cell_coordinate} — review_status={review}")

                            citation_match = re.search(r"\[(.*?),\s*pg\.\s*(\d+)\]", item.final_answer)
                            source_doc = "N/A"
                            page_num = "N/A"

                            if citation_match:
                                source_doc = citation_match.group(1)
                                page_num = f"pg. {citation_match.group(2)}"
                            elif item.sources:
                                source_doc = item.sources[0]

                            finalized_by = "User" if (item.edited_by_user or item.is_verified) else "AI"
                            review_label = (item.review_status or "pending").capitalize()

                            audit_ws.append([
                                f"{item.sheet_name}!{item.cell_coordinate}",
                                item.question,
                                item.final_answer if should_write else "(not approved)",
                                source_doc,
                                page_num,
                                item.confidence,
                                item.confidence_score if item.confidence_score is not None else "",
                                round(item.embedding_similarity_score, 4) if item.embedding_similarity_score is not None else "",
                                item.model_used or "",
                                review_label,
                                finalized_by,
                                item.answer_origin or "generated",
                            ])
                        except Exception as write_err:
                            print(f"⚠️ Error writing to {item.cell_coordinate}: {write_err}")
                            continue

                if any(item.confidence == "LOW" for item in answers):
                    print("⚠️ LOW CONFIDENCE ANSWERS PRESENT — REVIEW RECOMMENDED")

                # Save to fresh buffer while keeping source stream open
                output = BytesIO()
                wb.save(output)
                val = output.getvalue()
                output.close()
                return val
            finally:
                wb.close()
                in_buffer.close()
        except Exception as e:
            print(f"🛑 CRITICAL GENERATE ERROR: {e}")
            raise e

excel_agent = ExcelAgent()
