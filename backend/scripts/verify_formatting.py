from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Color
from io import BytesIO
from app.core.excel_agent import excel_agent
from app.models.schemas import QuestionItem

def verify_formatting():
    # 1. Create a styled Excel file
    wb = Workbook()
    ws = wb.active
    ws.title = "Formatting Test"
    
    # Header Style
    header_font = Font(bold=True, color="0000FF") # Blue
    ws["A1"].font = header_font
    ws["A1"] = "Project Name"
    ws["B1"] = "My Project"
    
    # Question
    ws["A3"] = "Describe your safety plan"
    
    # Answer Cell Style (Yellow Background)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws["B3"].fill = yellow_fill
    # B3 is empty initially
    
    # Save to bytes
    input_bytes = BytesIO()
    wb.save(input_bytes)
    input_bytes.seek(0)
    input_content = input_bytes.getvalue()
    
    print("✅ Created input file with styles.")
    
    # 2. Mock Analysis & Answer Injection
    print("🔄 Analyzing...")
    # We skip actual analysis to save time/tokens, just manually create the answer item
    # matching the cell B3
    
    answers = [
        QuestionItem(
            sheet_name="Formatting Test",
            cell_coordinate="B3",
            question="Describe your safety plan",
            ai_answer="We use guardrails.",
            final_answer="We use guardrails and nets.",
            confidence="LOW", # Trigger warning
            sources=["test_doc.pdf"],
            is_verified=False,
            edited_by_user=True
        )
    ]
    
    # 3. Generate Output
    print("💾 Generating Excel...")
    output_bytes_data = excel_agent.generate_excel(input_content, answers)
    
    # 4. Verify Output
    output_path = "formatted_test_output.xlsx"
    with open(output_path, "wb") as f:
        f.write(output_bytes_data)
        
    print(f"✅ Saved output to: {output_path}")
    
    # Reload and Check details
    wb_out = load_workbook(output_path)
    ws_out = wb_out["Formatting Test"]
    
    # Check Header Style
    cell_a1 = ws_out["A1"]
    if cell_a1.font.bold and cell_a1.font.color.rgb == "000000FF": # openpyxl uses ARGB
        print("✅ Header Style Preserved (Bold & Blue)")
    else:
        print(f"❌ Header Style LOST! Bold={cell_a1.font.bold}, Color={cell_a1.font.color.rgb}")

    # Check Answer Cell Style
    cell_b3 = ws_out["B3"]
    if cell_b3.fill.start_color.rgb == "00FFFF00":
         print("✅ Answer Cell Style Preserved (Yellow Background)")
    else:
         print(f"❌ Answer Cell Style LOST! Color={cell_b3.fill.start_color.rgb}")
         
    # Check Answer Value
    if cell_b3.value == "We use guardrails and nets.":
        print("✅ Answer Value Correct")
    else:
        print(f"❌ Answer Value WRONG: {cell_b3.value}")

    # Check Audit Sheet
    if "AI_Verification_Audit" in wb_out.sheetnames:
        print("✅ Audit Sheet Created")
        ws_audit = wb_out["AI_Verification_Audit"]
        # Check explicit columns
        headers = [c.value for c in ws_audit[1]]
        expected = ["Cell Reference", "Question", "Final Answer", "Source Document", "Confidence"]
        # We check if EXPECTED are IN metadata
        missing = [h for h in expected if h not in headers]
        if not missing:
             print("✅ Audit Columns Valid")
        else:
             print(f"❌ Missing Audit Columns: {missing}")
        
        # Check Content
        row_2 = [c.value for c in ws_audit[2]]
        print(f"   Audit Row 2: {row_2}")
        if "LOW" in row_2:
             print("✅ LOW Confidence Recorded")
    else:
        print("❌ Audit Sheet MISSING")

if __name__ == "__main__":
    verify_formatting()
