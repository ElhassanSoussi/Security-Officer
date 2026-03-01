from openpyxl import load_workbook
import os

def read_filled_excel():
    file_path = "test_questionnaire_FILLED.xlsx"
    if not os.path.exists(file_path):
        print(f"❌ File not found: {file_path}")
        return

    wb = load_workbook(file_path)
    
    print(f"📂 Reading: {file_path}")
    print("-" * 40)
    
    # 1. Read Main Sheet
    if "Safety Questionnaire" in wb.sheetnames:
        ws = wb["Safety Questionnaire"]
        print(f"\n📄 Sheet: {ws.title}")
        print(f"{'Question':<50} | {'Answer'}")
        print("-" * 80)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                q = row[0][:45] + "..." if len(row[0]) > 45 else row[0]
                a = row[1][:50] + "..." if row[1] and len(row[1]) > 50 else str(row[1])
                print(f"{q:<50} | {a}")
    
    # 2. Read Audit Sheet
    if "AI_Verification_Audit" in wb.sheetnames:
        ws = wb["AI_Verification_Audit"]
        print(f"\n🕵️ Sheet: {ws.title}")
        print(f"{'Cell':<10} | {'Confidence':<10} | {'Source':<30} | {'Answer Preview'}")
        print("-" * 80)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                # Row structure: Cell Ref, Question, Answer, Source, Page, Confidence, Finalized By
                cell_ref = row[0]
                confidence = row[5]
                source = row[3]
                answer = row[2][:30] + "..." if row[2] else ""
                print(f"{cell_ref:<10} | {confidence:<10} | {source:<30} | {answer}")

if __name__ == "__main__":
    read_filled_excel()
