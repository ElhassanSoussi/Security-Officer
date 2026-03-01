from openpyxl import Workbook
from openpyxl.styles import PatternFill
import os

def create_test_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Safety Questionnaire"
    
    # Headers
    ws["A1"] = "Question"
    ws["B1"] = "Answer"
    
    # Text Questions based on USER Request
    questions = [
        "Describe your fall protection plan",                # Should match 05005_Concrete.docx
        "Who is responsible for site safety?",               # Should match 05005_Concrete.docx
        "What is the company's Tax ID number?",              # Should be NOT FOUND IN LOCKER
        "List the safety protocols for concrete pouring."    # Relevant to Concrete doc
    ]
    
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    for i, q in enumerate(questions, start=2):
        ws[f"A{i}"] = q
        # B column left empty for AI to fill, but Styled
        ws[f"B{i}"].fill = yellow_fill
        
    # Save to current directory to avoid permission issues
    output_path = "test_questionnaire.xlsx"
    wb.save(output_path)
    print(f"✅ Created styled test file at: {output_path}")

if __name__ == "__main__":
    create_test_excel()
