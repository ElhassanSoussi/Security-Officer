from openpyxl import load_workbook
import sys

def verify_sprint6_final():
    print("🕵️ Verifying Sprint 6 Deliverables...")
    file_path = "test_questionnaire_FILLED.xlsx"
    
    try:
        wb = load_workbook(file_path)
    except Exception as e:
        print(f"❌ Failed to load workbook: {e}")
        sys.exit(1)
        
    ws = wb["Safety Questionnaire"]
    
    # Check 1: Answer Cell Style Preservation (Yellow Fill)
    # create_test_excel.py set B2, B3, B4, B5 to Yellow
    # run_extraction_test.py filled them with answers.
    # Did the fill survive?
    
    preserved_count = 0
    for row in range(2, 6):
        cell = ws[f"B{row}"]
        # openpyxl returns ARGB or Theme colors. If theme color, it might be tricky, but we set explicit RGB FFFF00
        # ARGB for Yellow is FFFF00 or 00FFFF00
        color = cell.fill.start_color.rgb
        # If it's a theme color, 'rgb' might be None or a theme index
        # PatternFill(start_color="FFFF00") usually results in rgb="00FFFF00"
        
        if color == "00FFFF00" or color == "FFFF00":
            preserved_count += 1
            
    if preserved_count == 4:
         print("✅ formatting: Answer Cell Yellow Background PRESERVED (4/4)")
    else:
         print(f"❌ formatting: Background Style LOST or CHANGED. Preserved: {preserved_count}/4")
         # print(f"Debug: {ws['B2'].fill.start_color.rgb}")

    # Check 2: Audit Sheet Existence
    if "AI_Verification_Audit" in wb.sheetnames:
        print("✅ audit: Sheet 'AI_Verification_Audit' EXISTS")
    else:
        print("❌ audit: Sheet MISSING")
        sys.exit(1)
        
    ws_audit = wb["AI_Verification_Audit"]
    
    # Check 3: Audit Columns
    headers = [c.value for c in ws_audit[1]]
    expected_cols = ["Cell Reference", "Question", "Final Answer", "Source Document", "Confidence"]
    missing = [c for c in expected_cols if c not in headers]
    
    if not missing:
        print("✅ audit: Columns VALID")
    else:
        print(f"❌ audit: Missing Columns: {missing}")
        
    # Check 4: Data Integrity
    
    # Row 2 (Question 1)
    row2 = [c.value for c in ws_audit[2]]
    print(f"   Row 2 Q: {row2[1]}")
    if row2[1] and "fall protection" in row2[1].lower():
         if row2[5] == "HIGH":
             print("✅ data: Q1 (Fall Protection) Confidence is HIGH")
         else:
             print(f"❌ data: Q1 Confidence is {row2[5]}")
    else:
         print(f"❌ data: Row 2 is not Fall Protection: {row2[1]}")

    # Row 4 (Question 3 - Tax ID)
    # Note: If Q2 was skipped before, this was the issue. Now Q2 should be Row 3.
    # So Tax ID should be Row 4.
    row4 = [c.value for c in ws_audit[4]]
    print(f"   Row 4 Q: {row4[1]}")
    if row4[1] and "tax id" in row4[1].lower():
        if row4[5] == "LOW":
             print("✅ data: Q3 (Tax ID) Confidence is LOW")
        else:
             print(f"❌ data: Q3 Confidence is {row4[5]}")
             
        if "NOT FOUND" in str(row4[2]):
             print("✅ data: Q3 Answer contains 'NOT FOUND'")
        else:
             print(f"❌ data: Q3 Answer missing 'NOT FOUND': {row4[2]}")
    else:
        print(f"❌ data: Row 4 is not Tax ID: {row4[1]}")

if __name__ == "__main__":
    verify_sprint6_final()
